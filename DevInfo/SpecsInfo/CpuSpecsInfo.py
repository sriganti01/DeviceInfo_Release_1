'''

@FileName : CpuSpecsInfo.py
@Author : Srinivas Ganti
@place : Hyderabad, 07 Jan 2024

@purpose : Class Contain Definition of functions
           for retrieving CPU Specifications from Mobile
'''

import logging
from SpecsInfo.MobileSpecsInfo import MobileSpecsInfo
from openpyxl.utils import get_column_letter
from Lib.FileSystemUtils import FileSystemUtils

log = logging.getLogger(__name__)

class ClusterInfo():

    def __init__(self):

        self.Number_Of_Cores = None
        self.Core_Architecture = None
        self.Cluster_Id = None

        self.Cores_List = None
        self.Cores_Offline = None
        self.Cores_Online = None
        self.Minimum_Frequency_Per_Core = None
        self.Maximum_Frequency_Per_Core = None
        self.Current_Frequency_Per_Core= None
        self.Cpu_Capacity_Per_Core = None
        self.Input_Boost_Freq = None
        self.Current_Governor_Per_Core = None
        self.Available_Frequencies_Per_Core = []
        self.Available_Governors__Per_Core = []

class CpuSpecsInfo(MobileSpecsInfo):

    '''
    Class for Software Specs  Object

    :param: None
    :return: None
    '''

    def __init__(self):
        '''
        @function:
            Initializes Software Specs Object

        @param: None
        @return: Non
        '''
        super().__init__()
        self.clusterlist = []
        self.clusterFoldersList = []
        self.nclusters = None
        self.devicefilepath= None

    def setup(self):
        super().setup()
        self.devicefilepath = '/sys'

    def grepInfo(self):
        self.addClustersToDict()
        return self.CpuSpecsInfoDict

    def cleanup(self):
        pass

    def setDeviceFilePath(self, filepath=None):
        if not filepath is None:
            self.devicefilepath = filepath

    def getDeviceFilePath(self):
        return self.devicefilepath
    def listDeviceContent(self, devfilepath=None):
        if devfilepath is None:
            devfilepath = self.getDeviceFilePath()
        self.command = self.ADBObj.getADBListCommand() + devfilepath
        self.output = self.executeCommandOnDevice(command=self.command)
        return self.output.splitlines()

    def readDeviceContent(self, devfilepath=None):
        if devfilepath is None:
            devfilepath = self.getDeviceFilePath()
        self.command = self.ADBObj.getADBReadContentCommand() + devfilepath
        self.output = self.executeCommandOnDevice(command=self.command)
        return self.output

    def getClustersFoldersList(self):
        self.clusterFoldersList = self.listDeviceContent(devfilepath='/sys/devices/system/cpu/cpufreq')

        return self.clusterFoldersList

    def getClusterCount(self):
        self.nclusters = len(self.getClustersFoldersList())

        return self.nclusters

    def convertToList(self, inputString = None, seperator = ' '):
        listData = inputString.split(seperator)
        return listData

    def getClusterInfo(self):
        self.getClusterCount()
        for cluster in range(0,self.nclusters):
            clusterInfo = ClusterInfo()
            clusterInfo.Cluster_Id = cluster
            filepath = '/sys/devices/system/cpu/cpufreq/' + str(self.clusterFoldersList[cluster])
            clusterInfo.Cores_List = self.convertToList(self.readDeviceContent(
                devfilepath=filepath +'/related_cpus'))
            clusterInfo.Number_Of_Cores = len(clusterInfo.Cores_List)
            clusterInfo.Minimum_Frequency_Per_Core = self.convertToList(self.readDeviceContent(
                devfilepath=filepath +'/scaling_min_freq'))
            clusterInfo.Maximum_Frequency_Per_Core = self.convertToList(self.readDeviceContent(
                devfilepath=filepath +'/scaling_max_freq'))
            clusterInfo.Current_Frequency_Per_Core = self.convertToList(self.readDeviceContent(
                devfilepath=filepath +'/scaling_cur_freq'))
            clusterInfo.Available_Frequencies_Per_Core = self.convertToList(self.readDeviceContent(
                devfilepath=filepath + '/scaling_available_frequencies'))
            clusterInfo.Available_Governors__Per_Core = self.convertToList(self.readDeviceContent(
                devfilepath=filepath + '/scaling_available_governors'))
            clusterInfo.Current_Governor_Per_Core = self.convertToList(self.readDeviceContent(
                devfilepath=filepath + '/scaling_governor'))
            clusterInfo.Cores_Offline = self.convertToList(self.readDeviceContent(
                devfilepath=filepath + '/../../offline'),seperator='-')
            clusterInfo.Cores_Online = self.convertToList(self.readDeviceContent(
                devfilepath=filepath + '/../../online'), seperator=',')
            clusterInfo.Input_Boost_Freq = self.convertToList(self.readDeviceContent(
                devfilepath=filepath + '/../../cpu_boost/input_boost_freq'))
            clusterInfo.Cpu_Capacity_Per_Core = self.convertToList(self.readDeviceContent(
                devfilepath=filepath + '/../../' +'cpu' + str(clusterInfo.Cores_List[0]) + '/cpu_capacity'))
            clusterInfo.Core_Architecture = self.convertToList(self.readDeviceContent(
                devfilepath='/proc/cpuinfo | grep Processor'),seperator=':')[-1].lstrip()

            log.info("Core Arch : {}".format(clusterInfo.Core_Architecture))

            self.clusterlist.append(clusterInfo)

        return self.clusterlist

    def addClustersToDict(self):
        self.getClusterInfo()
        for cluster  in self.clusterlist:
            self.updateDictionary(dictName=self.CpuSpecsInfoDict,
                                  key='Cluster' + str(cluster.Cluster_Id),
                                  value=cluster.__dict__)
        return self.CpuSpecsInfoDict

    def generateXLSXReport(self, xlsObj=None, wb=None,ws=None, dataDict=None):
        headers = list(dataDict.keys())
        headers.insert(0, "Clusters")

        for idx in range(0, len(headers)):
            cellref = ws.cell(row=2, column=idx + 2)
            ws.column_dimensions[get_column_letter(idx + 2)].width = 40
            cellref.style = xlsObj.getNamedStyle(stylename="headerRow")

            cellref.value = headers[idx]

        if isinstance(dataDict.get('Cluster0'), dict):
            dictkeys = list(dataDict.get('Cluster0').keys())
            for idx in range(0, len(dictkeys)):
                cellref = ws.cell(row=idx + 3, column=2)
                cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
                cellref.value = dictkeys[idx]

        headers = list(dataDict.keys())
        col_idx = 3
        for header in headers:
            row_idx = 3
            for datavalue in dataDict.get(header).values():
                cellref = ws.cell(row=row_idx, column=col_idx)
                cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
                charlist = ["[", "'", "]"]
                datavalue = FileSystemUtils.replaceChars(datavalue, charlist)
                cellref.value = str(datavalue)

                row_idx += 1
            col_idx += 1

        col_idx = 2
        row_idx = len(list(dataDict.get('Cluster0').keys())) +2
        for ctr in range(col_idx, col_idx+4):
            cellref = ws.cell(row=row_idx, column=ctr)
            cellref.style = xlsObj.getNamedStyle(stylename="lastRow")

