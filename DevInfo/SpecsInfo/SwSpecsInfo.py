'''

@FileName : SwSpecsInfo.py
@Author : Srinivas Ganti
@place : Hyderabad, 07 Jan 2024

@purpose : Class Contain Definition of functions
           for retrieving Software Specifications from Mobile
'''

import re
import logging
from Lib.ParserUtils import ParserUtils
from MobileSpecsInfo import MobileSpecsInfo
from openpyxl.utils import get_column_letter
from Lib.FileSystemUtils import FileSystemUtils

log = logging.getLogger(__name__)

class SwSpecsInfo(MobileSpecsInfo):

    '''
    Class for Software Specs  Object

    :param: None
    :return: None
    '''

    def __init(self):
        '''
        @function:
            Initializes Software Specs Object

        @param: None
        @return: None
        '''
        super().__init__()
        self.Adb_Version = None
        self.Kernel_Version = None
        self.Android_Version = None
        self.Open_GLes_Version = None
        self.Vendor_SDK_Version = None
        self.System_SDK_Version = None
        self.Android_SDK_Version = None
        self.Product_SDK_Version = None
        self.Device_Root_Status = None



    def grepInfo(self):
        self.getADBVersion()
        self.getAndroidVersion()
        self.getKernelVersion()
        self.getOpenGlesVersion()
        self.getDeviceRootStatus()
        self.getAndroidSDKVersion()
        self.getProductSDKVersion()
        self.getSystemSDKVersion()
        self.getVendorSDKVersion()

        return self.SwSpecsInfoDict

    def cleanup(self):
        pass

    def getADBVersion(self):
        '''
        @function: getADBVersion
            gets Current ADB Deamon Version Running on Device

        @param: None
        @return: ADB Daemon Version
        '''
        self.Adb_Version = self.executeCommandOnDevice(command=self.ADBObj.getADBVersionCommand())
        self.updateDictionary(dictName=self.SwSpecsInfoDict, key='Adb_Version', value=self.Adb_Version.splitlines())
        return self.Adb_Version

    def getAndroidVersion(self):
        '''
        @function: getAndroidVersion
            gets Current Android Version On Device

        @param: None
        @return: Current Android Version on Device
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.build.version.release '
        self.Android_Version = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.SwSpecsInfoDict, key='Android_Version', value=self.Android_Version)
        return self.Android_Version

    def getAndroidSDKVersion(self):
        '''
        @function: getAndroidSDKVersion
            gets Current Android SDK API Level Version On Device

        @param: None
        @return: Current Android SDK API Level Version on Device
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.build.version.sdk '
        self.Android_SDK_Version = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.SwSpecsInfoDict, key='Android_SDK_Version', value=self.Android_SDK_Version)
        return self.Android_SDK_Version

    def getVendorSDKVersion(self):
        '''
        @function: getVendorSDKVersion
            gets Vendor Current SDK API Level Version On Device

        @param: None
        @return: Vendor Current SDK API Level Version on Device
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.vendor.build.version.sdk '
        self.Vendor_SDK_Version = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.SwSpecsInfoDict, key='Vendor_SDK_Version', value=self.Vendor_SDK_Version)
        return self.Vendor_SDK_Version

    def getSystemSDKVersion(self):
        '''
        @function: getSystemSDKVersion
            gets System Current SDK API Level Version On Device

        @param: None
        @return: System Current SDK API Level Version on Device
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.system.build.version.sdk '
        self.System_SDK_Version = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.SwSpecsInfoDict, key='System_SDK_Version', value=self.System_SDK_Version)
        return self.System_SDK_Version

    def getProductSDKVersion(self):
        '''
        @function: getProductSDKVersion
            gets Product Current SDK API Level Version On Device

        @param: None
        @return: Product Current SDK API Level Version on Device
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.product.build.version.sdk '
        self.Product_SDK_Version = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.SwSpecsInfoDict, key='Product_SDK_Version', value=self.Product_SDK_Version)
        return self.Product_SDK_Version

    def getKernelVersion(self):
        '''
        @function: getKernelVersion
            gets Kernel Version On Device

        @param: None
        @return: Kernel Version on Device
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.kernel.version '
        self.Kernel_Version = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.SwSpecsInfoDict, key='Kernel_Version', value=self.Kernel_Version)
        return self.Kernel_Version

    def getOpenGlesVersion(self):
        self.command = self.ADBObj.getADBDumpsysCommand() + ' adreno'
        self.Open_GLes_Version = self.executeCommandOnDevice(command=self.command)
        pattern = re.compile(r'GLES: (?P<egl_version>.*)')
        rvalue = ParserUtils.parseDataViaRegex(pattern, self.Open_GLes_Version)
        self.updateDictionary(dictName=self.SwSpecsInfoDict, key='Open_GLes_Version', value=rvalue.get('egl_version'))
        return rvalue.get('egl_version')

    def getDeviceRootStatus(self):
        '''
        @function: getDeviceRootStatus
            gets Device Rootable Status

        @param: None
        @return: is Device Rootable or Not
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' persist.sys.is_root '
        self.Device_Root_Status = self.executeCommandOnDevice(command=self.command)
        if int(self.Device_Root_Status) == 0:
            self.Device_Root_Status = "Not Rooted"
        elif int(self.Device_Root_Status) == 1:
            self.Device_Root_Status = "Rooted"
        self.updateDictionary(dictName=self.SwSpecsInfoDict, key='Device_Root_Status',
                              value=self.Device_Root_Status)
        return self.Device_Root_Status


    def generateXLSXReport(self, xlsObj=None, wb=None,ws=None, dataDict=None):
        headers = []
        headers.insert(0, "Parameters")
        headers.insert(1, "Results")

        for idx in range(0, len(headers)):
            cellref = ws.cell(row=2, column=idx + 2)
            ws.column_dimensions[get_column_letter(idx + 2)].width = 40
            cellref.style = xlsObj.getNamedStyle(stylename="headerRow")

            cellref.value = headers[idx]


        dictkeys = list(dataDict.keys())
        for idx in range(0, len(dictkeys)):
            cellref = ws.cell(row=idx + 3, column=2)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            cellref.value = dictkeys[idx]

        #headers = list(dataDict.keys())
        col_idx = 3
        row_idx = 3
        for datavalue in dataDict.values():
            cellref = ws.cell(row=row_idx, column=col_idx)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            charlist = ["[", "'", "]"]
            datavalue = FileSystemUtils.replaceChars(datavalue, charlist)
            cellref.value = str(datavalue)
            row_idx += 1

        col_idx = 2
        row_idx = len(list(dataDict.keys())) + 2
        for ctr in range(col_idx, col_idx + 2):
            cellref = ws.cell(row=row_idx, column=ctr)
            cellref.style = xlsObj.getNamedStyle(stylename="lastRow")



