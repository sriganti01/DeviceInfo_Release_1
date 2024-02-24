'''

@FileName : HWSpecsInfo.py
@Author : Srinivas Ganti
@place : Hyderabad, 07 Jan 2024

@purpose : Class Contain Definition of functions
           for retrieving Hardware Specifications from Mobile
'''

import re
import logging
from Lib.ParserUtils import ParserUtils
from SpecsInfo.MobileSpecsInfo import MobileSpecsInfo
from openpyxl.utils import get_column_letter
from Lib.FileSystemUtils import FileSystemUtils

log = logging.getLogger(__name__)

class HWSpecsInfo(MobileSpecsInfo):

    '''
    Class for Hardware Specs  Object

    :param: None
    :return: None
    '''

    def __init(self):
        '''
        @function:
            Initializes Hardware Specs Object

        @param: None
        @return: None
        '''
        super().__init__()
        self.CameraSensorNames = []
        self.SocModel = None
        self.NFCChipType = None

        self.DeviceModemType = None
        self.UsbMtpDeviceType = None
        self.HardwareBaseBand = None
        self.DisplayPanelType = None
        self.NoOfCameraSensors = None
        self.ProductMaufacturer = None
        self.HardwareGPUPlatform = None
        self.DeviceRadioTypeList = None
        self.HardwareBoardPlatform = None
        self.ProductSocManufacturer = None
        self.PhysicalDeviceScreenSize = None
        self.PhysicalDeviceScreenDensity = None
        self.PhysicalDeviceScreenRotation = None
        self.PhysicalDeviceScreenMultiWindowConfig = None

    def grepInfo(self):
        self.CameraSensorNames = []
        self.getDeviceSerialNo()
        self.getHardwareBoardPlatform()
        self.getproductMaufacturer()
        self.getProductSocMaufacturer()
        self.getSocModel()
        self.getBaseBand()
        self.getDeviceModemType()
        self.getDeviceRadioTypeList()
        self.getDisplayPanelType()
        self.getGPUPlatform()
        self.getNFCChipType()
        self.getUSBMTPDeviceType()
        self.getNoOfCameras()
        self.getCameraSensorNames()
        self.getDevicePhysicalScreenSize()
        self.getDevicePhysicalScreenDensity()
        self.getDevicePhysicalScreenRotation()
        self.getDevicePhysicalScreenMultiWindowConfig()

        return self.HwSpecsInfoDict

    def cleanup(self):
        pass

    def getDeviceSerialNo(self):
        '''
        @function: getDeviceSerialNo
            gets Device Serial Number

        @param: None
        @return: Serial Number of Device
        '''
        self.deviceserialno = self.executeCommandOnDevice(command=self.ADBObj.getADBSerialNoCommand())
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='Deviceserialno', value=self.deviceserialno)
        return self.deviceserialno


    def getproductMaufacturer(self):
        '''
        @function: getproductMaufacturer
            gets Original Device Manufacturer Name

        @param: None
        @return: ODM Manufacturer Name
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.product.manufacturer '
        self.ProductMaufacturer = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='ProductManufacturer', value=self.ProductMaufacturer)
        return self.ProductMaufacturer

    def getProductSocMaufacturer(self):
        '''
        @function: getProductSocMaufacturer
            gets Original System-On-Chip Manufacturer Name

        @param: None
        @return: SOC Manufacturer Name
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.soc.manufacturer '
        self.ProductSocManufacturer = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='ProductSocManufacturer',
                              value=self.ProductSocManufacturer)
        return self.ProductSocManufacturer

    def getSocModel(self):
        '''
        @function: getSocModel
            gets Model of System-On-Chip

        @param: None
        @return: SOC Model Name
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.soc.model '
        self.SocModel = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='SocModel',
                              value=self.SocModel)
        return self.SocModel

    def getHardwareBoardPlatform(self):
        '''
        @function: getHardwareBoardPlatform
            gets Hardware Platform Name

        @param: None
        @return: Hardware Platform Name
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.board.platform '
        self.HardwareBoardPlatform = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='HardwareBoardPlatform',
                              value=self.HardwareBoardPlatform)
        return self.HardwareBoardPlatform

    def getBaseBand(self):
        '''
        @function: getBaseBand
            gets Radio BaseBand Name

        @param: None
        @return: Radio BaseBand Name
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.baseband '
        self.HardwareBaseBand = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='HardwareBaseBand',
                              value=self.HardwareBaseBand)
        return self.HardwareBaseBand

    def getGPUPlatform(self):
        '''
        @function: getGPUPlatform
            gets GPU Platform Name

        @param: None
        @return: GPU Platform Name
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.hardware.egl '
        self.HardwareGPUPlatform = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='HardwareGPUPlatform',
                              value=self.HardwareGPUPlatform)
        return self.HardwareGPUPlatform

    def getDisplayPanelType(self):
        '''
        @function: getDisplayPanelType
            gets Display Panel Type Name

        @param: None
        @return: Display Panel Type Name
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' | grep -i panel_type '
        self.DisplayPanelType = self.executeCommandOnDevice(command=self.command)
        #Need Parsing
        pattern = re.compile(r':\s+\[(?P<panel_type>.*)\]')
        rvalue = ParserUtils.parseDataViaRegex(pattern, self.DisplayPanelType)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='HardwareGPUPlatform',
                              value=rvalue.get('panel_type'))
        return rvalue.get('panel_type')

    def getNFCChipType(self):
        '''
        @function: getNFCChipType
            gets Near Field Communication Chip Type

        @param: None
        @return: nfc Chip Type
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' ro.hardware.nfc_nci '
        self.NFCChipType = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='NFCChipType',
                              value=self.NFCChipType)
        return self.NFCChipType

    def getNoOfCameras(self):
        '''
        @function: getNoOfCameras
            gets No Of Camera Sensors

        @param: None
        @return: number of Camera Sensors
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' persist.vendor.camera.sensor.number '
        self.NoOfCameraSensors = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='NoOfCameraSensors',
                              value=self.NoOfCameraSensors)
        return int(self.NoOfCameraSensors)

    def getCameraSensorNames(self):
        '''
        @function: getCameraSensorNames
            gets Camera Sensors Name

        @param: None
        @return: Camera Sensors Names
        '''
        for counter in range(0,self.getNoOfCameras()):
            self.command = self.ADBObj.getADBGetPropCommand() + ' persist.vendor.camera.sensor'+str(counter)
            self.CameraSensorNames.append(self.executeCommandOnDevice(command=self.command))
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='CameraSensorNames',
                              value=self.CameraSensorNames)
        return self.CameraSensorNames

    def getUSBMTPDeviceType(self):
        '''
        @function: getUSBMTPDeviceType
            gets Device Visible Type once connected via USB

        @param: None
        @return: Device Visible Type via USB
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' sys.usb.mtp.device_type '
        self.UsbMtpDeviceType = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='UsbMtpDeviceType',
                              value=self.UsbMtpDeviceType)
        return int(self.UsbMtpDeviceType)

    def getDeviceModemType(self):
        '''
        @function: getDeviceModemType
            gets Device Modem Type

        @param: None
        @return: Modem Type
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' persist.radio.multisim.config '
        self.DeviceModemType = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='DeviceModemType',
                              value=self.DeviceModemType)
        return self.DeviceModemType

    def getDeviceRadioTypeList(self):
        '''
        @function: getDeviceRadioTypeList
            gets Supported Radio Types List

        @param: None
        @return: Supported Radio Type List
        '''
        self.command = self.ADBObj.getADBGetPropCommand() + ' | grep -i radio.type.list '
        self.DeviceRadioTypeList = self.executeCommandOnDevice(command=self.command)
        pattern = re.compile(r':\s+\[(?P<radios_supported>.*)\]')
        rvalue = ParserUtils.parseDataViaRegex(pattern, self.DeviceRadioTypeList)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='DeviceRadioTypeList',
                              value=rvalue.get('radios_supported'))
        return rvalue.get('radios_supported')

    def getDevicePhysicalScreenSize(self):
        '''
        @function: getDevicePhysicalScreenSize
            gets Device Physical Screen Size

        @param: None
        @return: Device Physical Screen Size
        '''
        self.command = self.ADBObj.getADBWindowsManagerCommand() + ' size '
        self.physicalDeviceScreenSize = self.executeCommandOnDevice(command=self.command)
        pattern = re.compile(r':\s+(?P<screen_size>.*)')
        rvalue = ParserUtils.parseDataViaRegex(pattern,self.physicalDeviceScreenSize)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='PhysicalDeviceScreenSize',
                              value=rvalue.get('screen_size'))
        return rvalue.get('screen_size')

    def getDevicePhysicalScreenDensity(self):
        '''
        @function: getDevicePhysicalScreenDensity
            gets Device Physical Screen Density

        @param: None
        @return: Device Physical Screen Density
        '''
        self.command = self.ADBObj.getADBWindowsManagerCommand() + ' density '
        self.PhysicalDeviceScreenDensity = self.executeCommandOnDevice(command=self.command)
        #Need Parsing
        pattern = re.compile(r':\s+(?P<screen_density>.*)')
        rvalue = ParserUtils.parseDataViaRegex(pattern, self.PhysicalDeviceScreenDensity)

        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='PhysicalDeviceScreenDensity',
                              value=rvalue.get('screen_density'))
        return rvalue.get('screen_density')

    def getDevicePhysicalScreenRotation(self):
        '''
        @function: getDevicePhysicalScreenRotation
            gets Device Physical Screen Rotation

        @param: None
        @return: Device Physical Screen Rotation
        '''
        self.command = self.ADBObj.getADBWindowsManagerCommand() + ' user-rotation '
        self.PhysicalDeviceScreenRotation = self.executeCommandOnDevice(command=self.command)
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='PhysicalDeviceScreenRotation',
                              value=self.PhysicalDeviceScreenRotation)
        return self.PhysicalDeviceScreenRotation

    def getDevicePhysicalScreenMultiWindowConfig(self):
        '''
        @function: getDevicePhysicalScreenRotation
            gets Device Physical Screen Rotation

        @param: None
        @return: Device Physical Screen Rotation
        '''
        self.command = self.ADBObj.getADBWindowsManagerCommand() + ' get-multi-window-config '
        self.PhysicalDeviceScreenMultiWindowConfig = self.executeCommandOnDevice(command=self.command)
        #Need Parsing
        self.updateDictionary(dictName=self.HwSpecsInfoDict, key='PhysicalDeviceScreenMultiWindowConfig',
                              value=self.PhysicalDeviceScreenMultiWindowConfig.splitlines())
        return self.PhysicalDeviceScreenMultiWindowConfig

    def generateXLSXReport(self, xlsObj=None, wb=None,ws=None, dataDict=None):
        headers = []
        headers.insert(0, "Parameters")
        headers.insert(1, "Results")

        for idx in range(0, len(headers)):
            cellref = ws.cell(row=2, column=idx + 2)
            ws.column_dimensions[get_column_letter(idx + 2)].width = 40
            cellref.style = xlsObj.getNamedStyle(stylename="headerRow")

            cellref.value = headers[idx]


        dictkeys = list(dataDict.keys())
        for idx in range(0, len(dictkeys)):
            cellref = ws.cell(row=idx + 3, column=2)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            cellref.value = dictkeys[idx]

        #headers = list(dataDict.keys())
        col_idx = 3
        row_idx = 3
        for datavalue in dataDict.values():
            cellref = ws.cell(row=row_idx, column=col_idx)
            cellref.style = xlsObj.getNamedStyle(stylename="normalRow")
            charlist = ["[", "'", "]"]
            datavalue = FileSystemUtils.replaceChars(datavalue, charlist)
            cellref.value = str(datavalue)
            row_idx += 1

        col_idx = 2
        row_idx = len(list(dataDict.keys())) + 2
        for ctr in range(col_idx, col_idx + 2):
            cellref = ws.cell(row=row_idx, column=ctr)
            cellref.style = xlsObj.getNamedStyle(stylename="lastRow")

